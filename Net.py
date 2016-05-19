import numpy as np
import pdb
import math
import time
import gzip
import pickle
import matplotlib.cm as cm
import matplotlib.pyplot as plt
import matplotlib.image as mpimg
import os
from skimage import color
import openpyxl
import abc
import activation_functions
from openpyxl import load_workbook

#TODO:
#Implement the activation function interface 
#Implement run_in_reverse
#Turn monitoring on/off

error_list_training = []
error_list_validation = []

def load_data():
	"""Return the MNIST data as a tuple containing the training data,
	the validation data, and the test data.

	The ``training_data`` is returned as a tuple with two entries.
	The first entry contains the actual training images.  This is a
	numpy ndarray with 50,000 entries.  Each entry is, in turn, a
	numpy ndarray with 784 values, representing the 28 * 28 = 784
	pixels in a single MNIST image.

	The second entry in the ``training_data`` tuple is a numpy ndarray
	containing 50,000 entries.  Those entries are just the digit
	values (0...9) for the corresponding images contained in the first
	entry of the tuple.

	The ``validation_data`` and ``test_data`` are similar, except
	each contains only 10,000 images.

	This is a nice data format, but for use in neural networks it's
	helpful to modify the format of the ``training_data`` a little.
	That's done in the wrapper function ``load_data_wrapper()``, see
	below.
	"""
	f = gzip.open(os.getcwd() + '/mnist.pkl.gz', 'rb')
	training_data, validation_data, test_data = pickle.load(f, encoding='latin1')
	f.close()
	return (training_data, validation_data, test_data)


def load_data_wrapper():
	"""Return a tuple containing ``(training_data, validation_data,
	test_data)``. Based on ``load_data``, but the format is more
	convenient for use in our implementation of neural networks.

	In particular, ``training_data`` is a list containing 50,000
	2-tuples ``(x, y)``.  ``x`` is a 784-dimensional numpy.ndarray
	containing the input image.  ``y`` is a 10-dimensional
	numpy.ndarray representing the unit vector corresponding to the
	correct digit for ``x``.

	``validation_data`` and ``test_data`` are lists containing 10,000
	2-tuples ``(x, y)``.  In each case, ``x`` is a 784-dimensional
	numpy.ndarry containing the input image, and ``y`` is the
	corresponding classification, i.e., the digit values (integers)
	corresponding to ``x``.

	Obviously, this means we're using slightly different formats for
	the training data and the validation / test data.  These formats
	turn out to be the most convenient for use in our neural network
	code."""
	tr_d, va_d, te_d = load_data()
	training_inputs = [np.reshape(x, (784, 1)) for x in tr_d[0]]
	training_results = [vectorized_result(y) for y in tr_d[1]]
	training_data = np.array(list(zip(training_inputs, training_results)))
	validation_inputs = [np.reshape(x, (784, 1)) for x in va_d[0]]
	validation_results = [vectorized_result(y) for y in va_d[1]]
	validation_data = np.array(list(zip(validation_inputs, validation_results)))
	test_inputs = [np.reshape(x, (784, 1)) for x in te_d[0]]
	test_results = [vectorized_result(y) for y in te_d[1]]
	test_data = np.array(list(zip(test_inputs, test_results)))
	return (training_data, validation_data, test_data)


def vectorized_result(j):
	"""Return a 10-dimensional unit vector with a 1.0 in the jth
	position and zeroes elsewhere.  This is used to convert a digit
	(0...9) into a corresponding desired output from the neural
	network."""
	e = np.zeros((10, 1))
	e[j] = 1.0
	return e


# logging.disable(10)
class NeuralNetwork:
	def __init__(self, layers=None, activation_function=None, learningrate=0.0, lmb=0.0, sample_size=0, training_data=0, probability=0.5, momentum=0.0):
		#pdb.set_trace()
		if layers != None:
			self.make_layers(layers)
		self.learningrate = learningrate
		self.sample_size = sample_size
		self.lmb = lmb
		self.momentum = momentum
		self.dropout_probability = probability
		self.training_data = training_data
		if activation_function == None:
			self.activation_function = activation_functions.max()
		else:
			self.activation_function = activation_function

	def train(self, epochs):
		#pdb.set_trace()
		for j in range(0, epochs + 1):
			np.random.shuffle(self.training_data)
			print("Epoch number: " + str(j))
			for i in range(0, len(self.training_data), self.sample_size):
				#print("Epoch number: " + str(j) + "," + "Mini-batch number:" + str(i))
				tdarray = np.array(self.training_data[i:i + self.sample_size])
				if len(tdarray) < self.sample_size:
					break
				self.backprop(tdarray)
				# print("Epoch number " + str(j) + " complete")
			print("Accuracy: " + str(accuracy_test(self,t)))
			error_list_training.append(accuracy_test(self,t))
			error_list_validation.append(accuracy_test(self,v))
		plt.plot(error_list_training, 'r-', error_list_validation, 'b-')
		plt.show()

	def run(self, inp, do_dropout=False):
		#if do_dropout == True:
		#	pdb.set_trace()
		activations = []
		activations.append(inp)
		for i in range(0, len(self.weights)):
			bernoulli_matrix = np.ones(shape=(activations[i] @ np.transpose(self.weights[i])).shape)
			if 0 < i < (len(self.weights) - 1):
				if do_dropout == True:
					#pdb.set_trace()
					bernoulli_matrix = np.random.binomial(n=1,p=self.dropout_probability,size=(activations[i] @ np.transpose(self.weights[i])).shape)
				else:
					bernoulli_matrix = 1/(1-self.dropout_probability)
			activations.append(self.activation_function.activation_function((activations[i] @ np.transpose(self.weights[i]) + self.biases[i]))*bernoulli_matrix)
		return activations
				#Different things to try: multiply only max by bernoulli_matrix, multiply entire thing by matrix (so include biases),
				#--> and try not to include the final layer
				#activations.append(self.max((activations[i] @ np.transpose(self.weights[i]) + self.biases[i])))


	def run_inverse(self,inp):
		activations = []
		activation.append(inp)
		for i in range(len(self.weights)-1, -1, -1):
			activations.append(activations[i])

	def backprop(self, td):
		#pdb.set_trace()
		inputs = np.column_stack(np.array(td).transpose()[0]).transpose()
		outputs = np.column_stack(np.array(td).transpose()[1]).transpose()
		do_dropout = False
		if self.dropout_probability > 0 :
			do_dropout = True
		activations = self.run(inputs, do_dropout)
		# Below part is where bias gradient is calculated
		del_b = []
		del_b.append((1 / activations[-1][0].size) * (activations[-1] - outputs) * np.vectorize(self.activation_function.activation_prime)(
				activations[-1]))  # Bias change of output with Hadamard, hopefully?
		for i in range(len(self.layers) - 2, 0, -1):
			del_b.append(
					(del_b[len(self.layers) - 2 - i] @ self.weights[i]) * np.vectorize(self.activation_function.activation_prime)(activations[i]))
			# bias for hidden layers, no need to multiply by dropout matrix because the matrix is in activations(?) Also extend vs append
		# remember to divide sum by number of NONZERO entries
		# now for the weights
		del_b = list(reversed(del_b))
		# Multiply activation matrix by del_b matrix, hopefully the zeros in del_b will replace the Hadamard 1's and 0's matrix
		del_w = [((del_b[i]).transpose() @ activations[i])/float(self.sample_size) for i,_ in
				enumerate(del_b)]  # don't forget to divide by sample size, you just removed that
		del_b = [np.array(np.sum(del_b[i], axis=0)) * 1 / (self.sample_size) for i in
				 range(len(del_b))]  # Don't know if that last part is strictly necessary...
		# del_b = [(1/self.sample_size)*np.sum(del_b[i],axis=0) for i in range(len(del_b))]
		self.v_prime = [self.momentum*self.v_prime[i] -
				   self.learningrate*del_w[i] for i in
				   range(0,len(self.weights))]
		self.weights = [
			self.weights[i] + self.v_prime[i]  - (
			self.lmb / math.floor(len(self.training_data) / self.sample_size)) *
			self.weights[
				i] for i in range(0, len(self.weights))]
		self.biases = [self.biases[i] - (self.learningrate) * del_b[i] for i in range(len(self.biases))]
		# global error_list
		# error_list.append(self.mean_squared_error(activations[-1],outputs[-1]))
		# plt.plot(error_list)

	def show_results(self, t, full_output = False):
		for tset in t:
			output = self.run(np.array(tset[0]).transpose())[-1]
			maxindex = np.argmax(output)
			actual = np.argmax(tset[1])
			print("Output: " + str(maxindex) + "Actual: " + str(actual))
	
	def make_layers(self, layer):
		#pdb.set_trace()
		self.layers = layer
		self.weights = [.0001 * np.random.random_sample((layer[i + 1], layer[i])) for i in range(0, len(layer) - 1)]
		self.biases = [.0001*np.random.random_sample(layer[i]) for i in range(1, len(layer))]
		v_prime = [np.zeros((layer[i+1],layer[i])) for i in range(0, len(layer)-1)]
		self.v_prime = v_prime

	def mean_squared_error(self, o, e):
		return (.5 / o.shape[0]) * np.sum(np.sum((o - e) ** 2, axis=1) * 1 / self.sample_size)

	def save(self, link):
		wb = openpyxl.Workbook()
		wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
		for i in range(len(self.weights)):
			layer_name = "Weight matrix " + str(i+1)
			wb.create_sheet(index=i, title=layer_name)
			for j in range(self.weights[i].shape[0]):
				for k in range(self.weights[i].shape[1]):
					#pdb.set_trace()
					wb.get_sheet_by_name(layer_name).cell(row=(j+1), column=(k+1)).value = self.weights[i][j][k]
		wb.create_sheet(index=len(self.weights), title="Biases")
		for i in range(len(self.biases)):
			for j in range(len(self.biases[i])):
				wb.get_sheet_by_name("Biases").cell(row=(i+1), column=(j+1)).value = self.biases[i][j]
		wb.save(link)

	def load(self, link):
		wb = openpyxl.load_workbook(link, use_iterators=True)
		sheets = list(wb.worksheets)
		layers = [sheets[i].max_column for i in range(len(sheets) - 1)]
		layers.append(wb.worksheets[-2].max_row)
		self.make_layers(layers)
		#pdb.set_trace()
		self.weights = [np.array([[cell.value for cell in row] for row in sheets[i].iter_rows()]) for i in range(len(sheets)-1)]
		#weights = [[[float(sheets[i].cell(row=(j+1), column=(k+1)).value) 
		#	for k in range(self.weights[i].shape[1])] for j in range(self.weights[i].shape[0])] for i in range(len(sheets) - 1)]
		self.biases = [np.array([cell.value for cell in row if cell.value != None]) for row in sheets[-1].iter_rows()]

def accuracy_test(net, v):
	num_correct = 0
	for tset in v:
		output = net.run(np.array(tset[0]).transpose(), do_dropout=False)[-1]
		max_index = np.argmax(output)
		if max_index == np.argmax(tset[1]):
			num_correct += 1  # Misleading, but error actually tracks number correct
	return (float(num_correct) / float(len(v)))

def image_to_vector(file):
	vector = []
	img = color.rgb2gray(mpimg.imread(file))
	for i in range(img.shape[0]):
		for j in range(img.shape[1]):
			vector.append(float(1.0 - img[i][j]))
	return np.array(vector)


t, v, test = load_data_wrapper()
#plt.imshow(t[3][0].reshape((28,28)), cmap=cm.Greys_r)
#plt.show()
# print('The input' + str(t[0][0]))
# val_data = np.array(np.column_stack(v).transpose()[0]).transpose()
# print(val_data)
#print(v[0][0])
#toy_set = load_data_from_file('test.txt')
n = NeuralNetwork(layers=[784,30,10],learningrate=0.005, lmb=0.0, sample_size=20, training_data=t, 
	probability=0.5, momentum=0.5)
#print("Initial training set accuracy is: " + str(accuracy_test(n, t)))
#print("Initial validation set accuracy is: " + str(accuracy_test(n, v)))
#print("Starting...")d
#n.load("save_file.xlsx")
print("Accuracy rate of training set is: " + str(accuracy_test(n, t)))
n.train(epochs=100)
four_image = image_to_vector("4.png").reshape((784,1))
five_image = image_to_vector("5.png").reshape((784,1))
six_image = image_to_vector("6.png").reshape((784,1))
#plt.imshow(six_image.reshape((28,28)), cmap=cm.Greys_r)
#plt.show()
# print([np.transpos# e(bias) for bias in n.biases]
print("Accuracy rate of training set is: " + str(accuracy_test(n, t)))
print("Accuracy rate of validation set is:  " + str(accuracy_test(n, v)))
print("Accuracy rate of test set is: " + str(accuracy_test(n, test)))
four_result = np.argmax(n.run(four_image.reshape(1,784))[-1])
five_result = np.argmax(n.run(five_image.reshape(1,784))[-1])
six_result = np.argmax(n.run(six_image.reshape(1,784))[-1])
print(str(four_result))
print(str(five_result))
print(str(six_result))
#n.show_results(v)
#n.save("save_file.xlsx")